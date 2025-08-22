# -*- coding: utf-8 -*-
# Olive Young 모바일 브랜드 랭킹(1~100) 수집 -> 월별 시트 Excel 저장 -> Google Drive 업로드 -> Slack 알림(아마존 포맷)
# 규칙
# - 월 바뀌면 같은 파일 내 새 시트(예: "25년 9월")
# - 연도 바뀌면 파일명: "올리브영_브랜드_랭킹_YYYY.xlsx"
# - Slack 포맷: TOP 10 (전일 대비 변동 표시), 급상승/뉴브랜드/급하락/랭크 아웃
# - 급상승/급하락: 전체 Top100 중 ±10계단 이상, 각 최대 5개
# - 랭크 아웃: 전일 Top70에 있었으나 금일 Top100에 없음, 최대 5개

import os
import io
import re
import json
import logging
from datetime import datetime, timedelta, timezone

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from bs4 import BeautifulSoup

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload, MediaIoBaseDownload
from google.oauth2.credentials import Credentials as UserCredentials
from google.auth.transport.requests import Request as GoogleRequest

try:
    from playwright.sync_api import sync_playwright
    PLAYWRIGHT_AVAILABLE = True
except Exception:
    PLAYWRIGHT_AVAILABLE = False

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s: %(message)s")

SLACK_WEBHOOK = os.environ.get("SLACK_WEBHOOK_URL", "").strip()
GOOGLE_CLIENT_ID = os.environ.get("GOOGLE_CLIENT_ID", "").strip()
GOOGLE_CLIENT_SECRET = os.environ.get("GOOGLE_CLIENT_SECRET", "").strip()
GOOGLE_REFRESH_TOKEN = os.environ.get("GOOGLE_REFRESH_TOKEN", "").strip()
GDRIVE_FOLDER_ID = os.environ.get("GDRIVE_FOLDER_ID", "").strip()

TARGET_URL = "https://m.oliveyoung.co.kr/m/mtn?menu=ranking&tab=brands"
OUT_DIR = "out"

# -------------------- 공통 --------------------
def kst_now():
    return datetime.now(timezone.utc) + timedelta(hours=9)

def make_session_mobile():
    s = requests.Session()
    s.mount("https://", HTTPAdapter(max_retries=Retry(total=3, backoff_factor=1,
                                                     status_forcelist=[429,500,502,503,504])))
    s.headers.update({
        "User-Agent": ("Mozilla/5.0 (iPhone; CPU iPhone OS 16_0 like Mac OS X) "
                       "AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.0 Mobile/15E148 Safari/604.1"),
        "Accept-Language": "ko-KR,ko;q=0.9,en-US;q=0.8",
        "Referer": "https://m.oliveyoung.co.kr/",
    })
    return s

def send_slack(text: str):
    if not SLACK_WEBHOOK:
        logging.warning("SLACK_WEBHOOK_URL 미설정")
        return
    try:
        requests.post(SLACK_WEBHOOK, json={"text": text}, timeout=10)
    except Exception:
        logging.exception("Slack 전송 실패")

# -------------------- Drive --------------------
def drive_service():
    if not (GOOGLE_CLIENT_ID and GOOGLE_CLIENT_SECRET and GOOGLE_REFRESH_TOKEN):
        logging.warning("Google OAuth 환경변수 누락")
        return None
    try:
        creds = UserCredentials(
            None,
            refresh_token=GOOGLE_REFRESH_TOKEN,
            client_id=GOOGLE_CLIENT_ID,
            client_secret=GOOGLE_CLIENT_SECRET,
            token_uri="https://oauth2.googleapis.com/token",
            scopes=["https://www.googleapis.com/auth/drive.file"],
        )
        creds.refresh(GoogleRequest())
        return build("drive", "v3", credentials=creds, cache_discovery=False)
    except Exception:
        logging.exception("Drive 초기화 실패")
        return None

def find_file(service, folder_id, name):
    try:
        q = f"name='{name}' and mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'"
        if folder_id:
            q += f" and '{folder_id}' in parents"
        res = service.files().list(q=q, pageSize=1, fields="files(id,name)").execute()
        fs = res.get("files", [])
        return fs[0] if fs else None
    except Exception:
        logging.exception("find_file 실패")
        return None

def download_bytes(service, file_id):
    try:
        req = service.files().get_media(fileId=file_id)
        fh = io.BytesIO()
        downloader = MediaIoBaseDownload(fh, req)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        fh.seek(0)
        return fh.read()
    except Exception:
        logging.exception("다운로드 실패")
        return None

def upload_new(service, folder_id, filename, data: bytes):
    try:
        media = MediaIoBaseUpload(io.BytesIO(data),
                                  mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                  resumable=False)
        body = {"name": filename}
        if folder_id:
            body["parents"] = [folder_id]
        return service.files().create(body=body, media_body=media, fields="id,webViewLink").execute()
    except Exception:
        logging.exception("업로드 실패")
        return None

def update_file(service, file_id, data: bytes):
    try:
        media = MediaIoBaseUpload(io.BytesIO(data),
                                  mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                  resumable=False)
        return service.files().update(fileId=file_id, media_body=media).execute()
    except Exception:
        logging.exception("업데이트 실패")
        return None

# -------------------- 수집 --------------------
# footer/헤더 오탐 방지: "n위" 패턴과 함께 브랜드명 후보를 찾아 1~100 순으로 정렬
_RANK_PAT = re.compile(r"(^|\s)(\d{1,3})\s*위(?![^\n]*OUT)", re.M)

def is_brand_like(s: str) -> bool:
    if not s:
        return False
    s = s.strip()
    if len(s) < 1 or len(s) > 40:
        return False
    if any(x in s for x in ["브랜드", "더보기", "장바구니", "로그인", "이벤트", "고객센터",
                             "인스타그램", "페이스북", "유튜브", "대표전화", "채팅", "사업자", "개인정보"]):
        return False
    return True

def extract_rank_brand_from_li(li):
    """li 요소에서 (rank, brand) 추출. 실패 시 (None, None)"""
    try:
        txt = (li.inner_text() or "").strip()
        m = _RANK_PAT.search(txt)
        rank = None
        if m:
            rank = int(m.group(2))
            if not (1 <= rank <= 100):
                rank = None
        # 명시적 브랜드 셀렉터 우선
        for sel in [".brand_name", ".brand-name", ".brandNm", ".tx_brand", ".name", ".tit", ".title", "strong"]:
            el = li.query_selector(sel)
            if el:
                nm = (el.inner_text() or "").strip()
                nm = re.sub(r"\s{2,}", " ", nm)
                if is_brand_like(nm):
                    return rank, nm
        # 이미지 alt/aria-label 활용
        for sel in ["img[alt]", "[aria-label]"]:
            el = li.query_selector(sel)
            if el:
                val = (el.get_attribute("alt") or el.get_attribute("aria-label") or "").strip()
                # "메디힐 1위" 같은 경우 분리
                val = re.sub(r"\b\d{1,3}\s*위\b", "", val).strip()
                if is_brand_like(val):
                    return rank, val
        # 텍스트 라인에서 가장 브랜드스러운 한 줄 선택
        lines = [re.sub(r"\s{2,}", " ", x.strip()) for x in txt.split("\n") if x.strip()]
        lines = [re.sub(r"\b\d{1,3}\s*위\b", "", x).strip() for x in lines]
        cand = sorted(lines, key=lambda x: (-is_brand_like(x), -len(x)))
        for c in cand:
            if is_brand_like(c):
                return rank, c
        return rank, None
    except Exception:
        return None, None

def try_internal_api():
    # 만약 모바일 내부 API가 열려 있으면 사용(없으면 None 반환)
    session = make_session_mobile()
    candidates = [
        ("https://m.oliveyoung.co.kr/m/api/best/getBrandRankingList.do", {}),
        ("https://m.oliveyoung.co.kr/m/api/best/getBestBrandList.do", {}),
        ("https://m.oliveyoung.co.kr/m/api/best/brandRankList.do", {}),
    ]
    for url, params in candidates:
        try:
            r = session.get(url, params=params, timeout=10)
            if r.status_code != 200:
                continue
            data = r.json()
            for k in ["list", "rows", "items", "brandList", "data", "result"]:
                arr = data.get(k)
                if isinstance(arr, list) and arr:
                    out = []
                    for it in arr:
                        nm = it.get("brandNm") or it.get("brandName") or it.get("nm") or it.get("name")
                        rk = it.get("rank") or it.get("rk") or it.get("ord")
                        if nm and str(nm).strip():
                            out.append((int(rk) if rk else None, str(nm).strip()))
                    out = [(rk, nm) for rk, nm in out if nm]
                    if out:
                        out.sort(key=lambda x: (x[0] if x[0] else 9999))
                        return [nm for rk, nm in out][:100]
        except Exception:
            continue
    return None

def scrape_brands_top100():
    # 1) 내부 API 후보
    brands = try_internal_api()
    if brands and len(brands) >= 20:
        return brands[:100]

    # 2) Playwright로 모바일 페이지 렌더링 후 추출
    if not PLAYWRIGHT_AVAILABLE:
        logging.warning("Playwright 미설치")
        return []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True, args=["--no-sandbox", "--disable-dev-shm-usage"])
        device = p.devices.get("iPhone 12") or {}
        context = browser.new_context(**device, locale="ko-KR")
        page = context.new_page()
        page.goto(TARGET_URL, wait_until="networkidle", timeout=60000)
        # 랭킹 탭 확실히 로드
        page.wait_for_timeout(1500)

        # 스크롤 다운(무한 로딩 대비)
        for _ in range(6):
            page.evaluate("window.scrollBy(0, document.body.scrollHeight)")
            page.wait_for_timeout(500)

        # '브랜드'라는 단어가 포함된 랭킹 컨테이너 우선 탐색
        containers = page.query_selector_all("section, div, ul, ol")
        rank_items = []
        for c in containers:
            try:
                part_text = (c.inner_text() or "")
                if ("브랜드" in part_text) and ("랭킹" in part_text or "순위" in part_text):
                    lis = c.query_selector_all("li")
                    for li in lis:
                        rk, nm = extract_rank_brand_from_li(li)
                        if rk and nm:
                            rank_items.append((rk, nm))
            except Exception:
                pass

        # 보조: 페이지 전체에서 li 스캔(푸터 오탐 방지 위해 'n위' 필터 필수)
        if not rank_items:
            for li in page.query_selector_all("li"):
                rk, nm = extract_rank_brand_from_li(li)
                if rk and nm:
                    rank_items.append((rk, nm))

        browser.close()

    # 랭크 중복 제거 & 정렬
    by_rank = {}
    for rk, nm in rank_items:
        if 1 <= rk <= 100 and nm and is_brand_like(nm):
            if rk not in by_rank:
                by_rank[rk] = nm
    brands = [by_rank.get(i, "") for i in range(1, 101)]
    brands = [b for b in brands if b]
    return brands[:100]

# -------------------- 엑셀 --------------------
def month_sheet(dt: datetime) -> str:
    return f"{dt.year % 100}년 {dt.month}월"

def file_name(dt: datetime) -> str:
    return f"올리브영_브랜드_랭킹_{dt.year}.xlsx"

def ensure_header(ws):
    if ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None:
        ws.append(["날짜"] + [f"{i}위" for i in range(1, 101)])
        ws.column_dimensions["A"].width = 14
        for c in range(2, 102):
            ws.column_dimensions[get_column_letter(c)].width = 12

def write_today(ws, dt: datetime, brands):
    d = dt.date().isoformat()
    row = None
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, 1).value) == d:
            row = r
            break
    if row is None:
        row = ws.max_row + 1
    ws.cell(row=row, column=1, value=d)
    for i in range(100):
        ws.cell(row=row, column=2 + i, value=(brands[i] if i < len(brands) else ""))

def read_prev_map(wb, now_dt: datetime):
    """전일(또는 직전) 랭크 맵 brand->rank. 같은 달에서 전일 없으면 이전 달 시트의 마지막 행 사용."""
    def row_to_map(ws, row):
        m = {}
        for i in range(1, 101):
            name = (ws.cell(row=row, column=1 + i).value or "").strip() if ws.cell(row=row, column=1 + i).value else ""
            if name:
                m[name] = i
        return m

    # 후보: 같은 시트에서 오늘 이전 날짜 중 가장 최근
    cur = month_sheet(now_dt)
    cand = []
    if cur in wb.sheetnames:
        ws = wb[cur]
        for r in range(2, ws.max_row + 1):
            try:
                d = str(ws.cell(r, 1).value)
                if d and d < now_dt.date().isoformat():
                    cand.append((ws, r))
            except Exception:
                continue
    # 이전 달 시트(있으면) 마지막 행
    prev_month_dt = (now_dt.replace(day=1) - timedelta(days=1))
    prev_name = month_sheet(prev_month_dt)
    if prev_name in wb.sheetnames:
        ws2 = wb[prev_name]
        if ws2.max_row >= 2:
            cand.append((ws2, ws2.max_row))

    if not cand:
        return {}
    # 가장 최근 행 선택(마지막 후보)
    ws, r = cand[-1]
    return row_to_map(ws, r)

# -------------------- 슬랙 메시지 --------------------
def format_delta(today_rank, prev_rank):
    if prev_rank is None:
        return "NEW"
    diff = prev_rank - today_rank
    if diff > 0:
        return f"↑{diff}"
    elif diff < 0:
        return f"↓{abs(diff)}"
    else:
        return "—"

def build_sections(today_list, prev_map):
    # today_map
    today_map = {b: i+1 for i, b in enumerate(today_list)}
    # TOP10
    top10_lines = []
    for i in range(min(10, len(today_list))):
        b = today_list[i]
        pr = prev_map.get(b)
        delta = format_delta(i+1, pr)
        top10_lines.append(f"{i+1}. ({delta}) {b}")

    # 급상승/급하락
    ups, downs = [], []
    for b, tr in today_map.items():
        pr = prev_map.get(b)
        if pr is None:
            continue
        diff = pr - tr  # +면 상승
        if diff >= 10:
            ups.append((diff, b, pr, tr))
        elif diff <= -10:
            downs.append((abs(diff), b, pr, tr))
    ups.sort(key=lambda x: (-x[0], x[3]))      # 큰 폭 우선
    downs.sort(key=lambda x: (-x[0], x[3]))    # 큰 폭 우선
    ups = [f"- {b} {pr}위 → {tr}위 (↑{d})" for d, b, pr, tr in ups[:5]]
    downs = [f"- {b} {pr}위 → {tr}위 (↓{d})" for d, b, pr, tr in downs[:5]]

    # 뉴브랜드(IN)
    new_in = [b for b in today_list if b not in prev_map]
    new_in_lines = [f"- {b} NEW → {today_map[b]}위" for b in new_in[:5]]

    # 랭크 아웃(전일 <=70 이었는데 금일 Top100에서 사라짐)
    prev_top70 = [(b, r) for b, r in prev_map.items() if r <= 70]
    outs = []
    today_set = set(today_list)
    for b, r in sorted(prev_top70, key=lambda x: x[1]):
        if b not in today_set:
            outs.append((b, r))
    out_lines = [f"- {b} {r}위 → OUT" for b, r in outs[:5]]

    inout_summary = f"{len(new_in)}개 IN, {len(outs)}개 OUT"

    return top10_lines, ups, new_in_lines, downs, out_lines, inout_summary

# -------------------- 메인 --------------------
def main():
    now = kst_now()
    logging.info("브랜드 랭킹 수집 시작")

    brands = scrape_brands_top100()
    if len(brands) < 20:
        send_slack("❌ *올리브영 모바일 브랜드 랭킹* 수집 실패 (데이터 부족)")
        return 1

    # Excel 파일 준비
    svc = drive_service()
    fname = file_name(now)
    meta = find_file(svc, GDRIVE_FOLDER_ID, fname)

    wb = None
    if meta:
        data = download_bytes(svc, meta["id"])
        if data:
            wb = load_workbook(io.BytesIO(data))
    if wb is None:
        wb = Workbook()

    sheet = month_sheet(now)
    if sheet in wb.sheetnames:
        ws = wb[sheet]
    else:
        ws = wb.create_sheet(title=sheet)
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        try:
            wb.remove(wb["Sheet"])
        except Exception:
            pass

    ensure_header(ws)
    prev_map = read_prev_map(wb, now)
    write_today(ws, now, brands)

    # 저장 & 업로드
    os.makedirs(OUT_DIR, exist_ok=True)
    path = os.path.join(OUT_DIR, fname)
    wb.save(path)
    with open(path, "rb") as f:
        xbytes = f.read()

    view_link = None
    if meta:
        update_file(svc, meta["id"], xbytes)
        try:
            view_link = svc.files().get(fileId=meta["id"], fields="webViewLink").execute().get("webViewLink")
        except Exception:
            pass
    else:
        created = upload_new(svc, GDRIVE_FOLDER_ID, fname, xbytes)
        view_link = (created or {}).get("webViewLink")

    # 슬랙 메시지
    top10, ups, newins, downs, outs, inout_summary = build_sections(brands, prev_map)
    msg = [
        f"*올리브영 모바일 브랜드 랭킹 100* — {now.strftime('%Y-%m-%d')}",
        f"- 월 시트: `{sheet}` / 파일: `{fname}`",
        "",
        "*TOP 10*",
        *top10,
        "",
        "🔥 *급상승* (10계단↑, 최대 5개)" if ups else "🔥 *급상승*: 해당 없음",
        *(ups if ups else []),
        "",
        "🆕 *뉴브랜드* (오늘 Top100 신규, 최대 5개)" if newins else "🆕 *뉴브랜드*: 해당 없음",
        *(newins if newins else []),
        "",
        "📉 *급하락* (10계단↓, 최대 5개)" if downs else "📉 *급하락*: 해당 없음",
        *(downs if downs else []),
        "",
        "⬅️ *랭크 아웃* (전일 Top70 → 금일 OUT, 최대 5개)" if outs else "⬅️ *랭크 아웃*: 해당 없음",
        *(outs if outs else []),
        "",
        f"➡️ 랭크 인&아웃 요약: {inout_summary}",
    ]
    if view_link:
        msg.append(f"\n<{view_link}|Google Drive에서 열기>")
    send_slack("\n".join(msg))

    logging.info("완료")
    return 0

if __name__ == "__main__":
    raise SystemExit(main())
